from flask import Blueprint, render_template, request, jsonify, send_file
import mysql.connector
from decimal import Decimal, InvalidOperation
import csv
import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from notifications import get_conteggio_codici, controlla_e_notifica, SOGLIA
import os

admin_bp = Blueprint('admin', __name__)

# Configurazione database (deve corrispondere a quella in app.py)
DB_CONFIG = {
    'host': os.environ.get('MYSQLHOST', 'localhost'),
    'user': os.environ.get('MYSQLUSER', 'root'),
    'password': os.environ.get('MYSQLPASSWORD', '12345678'),
    'database': os.environ.get('MYSQLDATABASE', 'CDD_YM'),
    'port': int(os.environ.get('MYSQLPORT', 3306))
}

def get_db_connection():
    return mysql.connector.connect(**DB_CONFIG)


@admin_bp.route('/admin')
def admin():
    return render_template('admin.html')


@admin_bp.route('/carica', methods=['POST'])
def carica_codici():
    """
    Input: multipart/form-data con chiave 'file' (CSV)
    Colonne attese: CodiceID, Tipo, Importo, Edizione
    Output: { "inseriti": int, "duplicati": int, "errori": [str, ...] }
    """
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Nessun file fornito'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'File non selezionato'}), 400
        if not file.filename.lower().endswith('.csv'):
            return jsonify({'error': 'Solo file CSV (.csv) sono accettati'}), 400

        contenuto = file.read().decode('utf-8')
        reader = csv.DictReader(io.StringIO(contenuto))

        righe_valide = []
        errori = []

        for i, riga in enumerate(reader, start=2):
            codice   = riga.get('CodiceID', '').strip()
            tipo     = riga.get('Tipo', '').strip().upper()
            importo  = riga.get('Importo', '').strip()
            edizione = riga.get('Edizione', '').strip()

            if not codice:
                errori.append(f'Riga {i}: CodiceID mancante')
                continue
            if len(codice) > 64:
                errori.append(f'Riga {i}: CodiceID troppo lungo ({len(codice)} caratteri)')
                continue
            if tipo not in ('CDD', 'YM'):
                errori.append(f'Riga {i}: Tipo "{tipo}" non valido (usa CDD o YM)')
                continue
            if not edizione:
                errori.append(f'Riga {i}: Edizione mancante')
                continue
            try:
                importo_dec = Decimal(importo)
                if importo_dec <= 0:
                    raise ValueError
            except (InvalidOperation, ValueError):
                errori.append(f'Riga {i}: Importo "{importo}" non valido')
                continue

            righe_valide.append((codice, tipo, importo_dec, edizione))

        if not righe_valide:
            return jsonify({'inseriti': 0, 'duplicati': 0, 'errori': errori})

        inseriti = 0
        duplicati = 0

        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            sql = "INSERT IGNORE INTO Codici (CodiceID, Tipo, Importo, Edizione, StatoCodice) VALUES (%s, %s, %s, %s, 'Disponibile')"
            for riga in righe_valide:
                cursor.execute(sql, riga)
                if cursor.rowcount == 1:
                    inseriti += 1
                else:
                    duplicati += 1
            conn.commit()
        except Exception as e:
            conn.rollback()
            cursor.close()
            conn.close()
            print(f'Errore DB upload: {e}')
            return jsonify({'error': 'Errore durante il salvataggio nel database'}), 500

        cursor.close()
        conn.close()

        return jsonify({'inseriti': inseriti, 'duplicati': duplicati, 'errori': errori})

    except Exception as e:
        print(f'Errore upload: {e}')
        return jsonify({'error': 'Errore interno del server'}), 500


@admin_bp.route('/admin/stato-codici')
def stato_codici():
    """
    Restituisce i conteggi dei codici disponibili per Tipo, Edizione e Importo,
    con lo stato attivo/disabilitato per campagna e taglio.
    Output: { "codici": [{ tipo, edizione, importo, disponibili, campagna_attiva, taglio_attivo }], "soglia": int }
    """
    try:
        conteggi = get_conteggio_codici()

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            "SELECT chiave, valore FROM Configurazione WHERE chiave LIKE 'disabilitato_%'"
        )
        stati = {row[0]: row[1] for row in cursor.fetchall()}
        cursor.close()
        conn.close()

        for c in conteggi:
            chiave_campagna = f"disabilitato_{c['tipo']}_{c['edizione']}"
            chiave_taglio = f"disabilitato_{c['tipo']}_{c['edizione']}_{c['importo']:.2f}"
            c['campagna_attiva'] = stati.get(chiave_campagna) != '1'
            c['taglio_attivo'] = stati.get(chiave_taglio) != '1'

        return jsonify({'codici': conteggi, 'soglia': SOGLIA})
    except Exception as e:
        print(f'Errore monitoraggio: {e}')
        return jsonify({'error': 'Errore nel recupero dei dati'}), 500


@admin_bp.route('/admin/toggle-campagna', methods=['POST'])
def toggle_campagna():
    """
    Abilita/disabilita una campagna (Tipo + Edizione).
    Input: { "tipo": "CDD", "edizione": "2025" }
    Output: { "attivo": bool }
    """
    try:
        data = request.get_json()
        tipo = data.get('tipo', '').upper().strip()
        edizione = data.get('edizione', '').strip()

        if tipo not in ('CDD', 'YM') or not edizione:
            return jsonify({'error': 'Parametri non validi'}), 400

        chiave = f'disabilitato_{tipo}_{edizione}'
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT valore FROM Configurazione WHERE chiave = %s", (chiave,))
        row = cursor.fetchone()
        nuovo_valore = '0' if (row and row[0] == '1') else '1'
        cursor.execute(
            "INSERT INTO Configurazione (chiave, valore) VALUES (%s, %s) "
            "ON DUPLICATE KEY UPDATE valore = %s",
            (chiave, nuovo_valore, nuovo_valore)
        )
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'attivo': nuovo_valore != '1'})
    except Exception as e:
        print(f'Errore toggle campagna: {e}')
        return jsonify({'error': 'Errore interno del server'}), 500


@admin_bp.route('/admin/toggle-taglio', methods=['POST'])
def toggle_taglio():
    """
    Abilita/disabilita un taglio specifico (Tipo + Edizione + Importo).
    Input: { "tipo": "CDD", "edizione": "2025", "importo": "25.00" }
    Output: { "attivo": bool }
    """
    try:
        data = request.get_json()
        tipo = data.get('tipo', '').upper().strip()
        edizione = data.get('edizione', '').strip()
        importo_str = data.get('importo', '').strip()

        if tipo not in ('CDD', 'YM') or not edizione:
            return jsonify({'error': 'Parametri non validi'}), 400

        try:
            importo_dec = Decimal(importo_str)
            if importo_dec <= 0:
                raise ValueError
        except (InvalidOperation, ValueError):
            return jsonify({'error': 'Importo non valido'}), 400

        chiave = f'disabilitato_{tipo}_{edizione}_{float(importo_dec):.2f}'
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT valore FROM Configurazione WHERE chiave = %s", (chiave,))
        row = cursor.fetchone()
        nuovo_valore = '0' if (row and row[0] == '1') else '1'
        cursor.execute(
            "INSERT INTO Configurazione (chiave, valore) VALUES (%s, %s) "
            "ON DUPLICATE KEY UPDATE valore = %s",
            (chiave, nuovo_valore, nuovo_valore)
        )
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'attivo': nuovo_valore != '1'})
    except Exception as e:
        print(f'Errore toggle taglio: {e}')
        return jsonify({'error': 'Errore interno del server'}), 500


@admin_bp.route('/admin/stato-sistema')
def stato_sistema():
    """
    Restituisce lo stato attuale della distribuzione.
    Output: { "attivo": bool }
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT valore FROM Configurazione WHERE chiave = 'distribuzione_attiva'")
        row = cursor.fetchone()
        cursor.close()
        conn.close()
        return jsonify({'attivo': row is not None and row[0] == '1'})
    except Exception as e:
        print(f'Errore stato sistema: {e}')
        return jsonify({'error': 'Errore interno del server'}), 500


@admin_bp.route('/admin/toggle-sistema', methods=['POST'])
def toggle_sistema():
    """
    Inverte lo stato della distribuzione (attivo ↔ disattivato).
    Output: { "attivo": bool }
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT valore FROM Configurazione WHERE chiave = 'distribuzione_attiva'")
        row = cursor.fetchone()
        nuovo_valore = '0' if (row and row[0] == '1') else '1'
        cursor.execute(
            "UPDATE Configurazione SET valore = %s WHERE chiave = 'distribuzione_attiva'",
            (nuovo_valore,)
        )
        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'attivo': nuovo_valore == '1'})
    except Exception as e:
        print(f'Errore toggle sistema: {e}')
        return jsonify({'error': 'Errore interno del server'}), 500


@admin_bp.route('/admin/campagne')
def campagne():
    """
    Restituisce la lista delle campagne presenti nel database con i conteggi dei codici.
    Output: { "campagne": [{ "tipo", "edizione", "disponibili", "usati" }] }
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT Tipo, Edizione,
                SUM(CASE WHEN StatoCodice = 'Disponibile' THEN 1 ELSE 0 END),
                SUM(CASE WHEN StatoCodice = 'Usato' THEN 1 ELSE 0 END)
            FROM Codici
            GROUP BY Tipo, Edizione
            ORDER BY Tipo, Edizione
        """)
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        result = [{'tipo': r[0], 'edizione': r[1], 'disponibili': int(r[2]), 'usati': int(r[3])} for r in rows]
        return jsonify({'campagne': result})
    except Exception as e:
        print(f'Errore campagne: {e}')
        return jsonify({'error': 'Errore nel recupero delle campagne'}), 500


@admin_bp.route('/admin/elimina-campagna', methods=['POST'])
def elimina_campagna():
    """
    Elimina tutti i codici Disponibili di una campagna (Tipo + Edizione).
    I codici già assegnati (Usato) non vengono toccati.
    Input: { "tipo": "CDD", "edizione": "2024" }
    Output: { "eliminati": int }
    """
    try:
        data = request.get_json()
        tipo = data.get('tipo', '').upper().strip()
        edizione = data.get('edizione', '').strip()

        if not tipo or not edizione:
            return jsonify({'error': 'Parametri mancanti'}), 400

        conn = get_db_connection()
        cursor = conn.cursor()

        # Verifica che la campagna esista
        cursor.execute(
            "SELECT COUNT(*) FROM Codici WHERE Tipo = %s AND Edizione = %s",
            (tipo, edizione)
        )
        if cursor.fetchone()[0] == 0:
            cursor.close()
            conn.close()
            return jsonify({'error': 'Campagna non trovata'}), 404

        # Elimina solo i codici Disponibili
        cursor.execute(
            "DELETE FROM Codici WHERE Tipo = %s AND Edizione = %s AND StatoCodice = 'Disponibile'",
            (tipo, edizione)
        )
        eliminati = cursor.rowcount

        # Rimuove le chiavi di configurazione associate a questa campagna
        cursor.execute(
            "DELETE FROM Configurazione WHERE chiave LIKE %s",
            (f'disabilitato_{tipo}_{edizione}%',)
        )

        conn.commit()
        cursor.close()
        conn.close()
        return jsonify({'eliminati': eliminati})
    except Exception as e:
        print(f'Errore elimina campagna: {e}')
        return jsonify({'error': 'Errore interno del server'}), 500


def _stile_intestazione(ws, headers, colore_hex='FFF3E8'):
    """Applica stile alle intestazioni e imposta la larghezza delle colonne."""
    fill = PatternFill(fill_type='solid', fgColor=colore_hex)
    font = Font(bold=True, size=11)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # Larghezza automatica approssimativa
        ws.column_dimensions[cell.column_letter].width = max(len(header) + 4, 14)
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 20


@admin_bp.route('/admin/export/codici')
def export_codici():
    """
    Esporta tutti i codici con dettagli ordine associato in formato Excel.
    Colonne: CodiceID, Tipo, Edizione, Importo, Stato, ID Ordine,
             Numero Ordine, Contatto, Motivazione, Dettaglio, Data
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT
                c.CodiceID, c.Tipo, c.Edizione, c.Importo, c.StatoCodice,
                o.IdentificativoOrdine, o.Ordine, o.Contatto,
                o.Motivazione, o.MotivazioneDettaglio, o.DataUtilizzo
            FROM Codici c
            LEFT JOIN Ordini o ON o.IdentificativoOrdine = c.IdentificativoOrdine
            ORDER BY c.Tipo, c.Edizione, c.StatoCodice, c.Importo DESC, c.CodiceID
        """)
        rows = cursor.fetchall()
        cursor.close()
        conn.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Codici'

        headers = ['CodiceID', 'Tipo', 'Edizione', 'Importo (€)', 'Stato',
                   'ID Ordine Interno', 'Numero Ordine', 'Contatto',
                   'Motivazione', 'Dettaglio Motivazione', 'Data Assegnazione']
        _stile_intestazione(ws, headers)

        for row in rows:
            ws.append([
                row[0], row[1], row[2], float(row[3]) if row[3] else None,
                row[4], row[5], row[6], row[7], row[8], row[9],
                row[10].strftime('%d/%m/%Y') if row[10] else None
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f'codici_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return send_file(output,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         download_name=filename)
    except Exception as e:
        print(f'Errore export codici: {e}')
        return jsonify({'error': 'Errore durante la generazione del file'}), 500


@admin_bp.route('/admin/export/ordini')
def export_ordini():
    """
    Esporta tutti gli ordini con contatto, motivazione, totale e lista codici.
    Colonne: ID Ordine, Numero Ordine, Contatto, Motivazione, Dettaglio,
             Data, N. Voucher, Totale €, Codici Assegnati
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT
                o.IdentificativoOrdine, o.Ordine, o.Contatto,
                o.Motivazione, o.MotivazioneDettaglio, o.DataUtilizzo,
                COUNT(c.CodiceID) AS NumeroVoucher,
                SUM(c.Importo) AS TotaleEuro,
                GROUP_CONCAT(c.CodiceID ORDER BY c.CodiceID SEPARATOR ', ') AS Codici
            FROM Ordini o
            LEFT JOIN Codici c ON c.IdentificativoOrdine = o.IdentificativoOrdine
            GROUP BY o.IdentificativoOrdine, o.Ordine, o.Contatto,
                     o.Motivazione, o.MotivazioneDettaglio, o.DataUtilizzo
            ORDER BY o.DataUtilizzo DESC, o.IdentificativoOrdine DESC
        """)
        rows = cursor.fetchall()
        cursor.close()
        conn.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Ordini'

        headers = ['ID Ordine', 'Numero Ordine', 'Contatto', 'Motivazione',
                   'Dettaglio Motivazione', 'Data', 'N. Voucher', 'Totale (€)', 'Codici Assegnati']
        _stile_intestazione(ws, headers)

        for row in rows:
            ws.append([
                row[0], row[1], row[2], row[3], row[4],
                row[5].strftime('%d/%m/%Y') if row[5] else None,
                int(row[6]) if row[6] else 0,
                float(row[7]) if row[7] else 0.0,
                row[8]
            ])
        # Colonna "Codici Assegnati" più larga
        ws.column_dimensions['I'].width = 60

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f'ordini_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return send_file(output,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         download_name=filename)
    except Exception as e:
        print(f'Errore export ordini: {e}')
        return jsonify({'error': 'Errore durante la generazione del file'}), 500


@admin_bp.route('/admin/export/riepilogo')
def export_riepilogo():
    """
    Esporta il riepilogo aggregato per campagna (Tipo + Edizione + Importo).
    Colonne: Tipo, Edizione, Taglio (€), Disponibili, Usati, Totale
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT
                Tipo, Edizione, Importo,
                SUM(CASE WHEN StatoCodice = 'Disponibile' THEN 1 ELSE 0 END) AS Disponibili,
                SUM(CASE WHEN StatoCodice = 'Usato' THEN 1 ELSE 0 END) AS Usati,
                COUNT(*) AS Totale
            FROM Codici
            GROUP BY Tipo, Edizione, Importo
            ORDER BY Tipo, Edizione, Importo DESC
        """)
        rows = cursor.fetchall()
        cursor.close()
        conn.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Riepilogo Campagne'

        headers = ['Tipo', 'Edizione', 'Taglio (€)', 'Disponibili', 'Usati', 'Totale']
        _stile_intestazione(ws, headers)

        for row in rows:
            ws.append([row[0], row[1], float(row[2]), int(row[3]), int(row[4]), int(row[5])])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f'riepilogo_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return send_file(output,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         download_name=filename)
    except Exception as e:
        print(f'Errore export riepilogo: {e}')
        return jsonify({'error': 'Errore durante la generazione del file'}), 500


@admin_bp.route('/admin/invia-notifica', methods=['POST'])
def invia_notifica():
    """
    Controlla i livelli e invia l'email di notifica se qualcuno è sotto soglia.
    Output: { "sotto_soglia": [...], "email_inviata": bool }
    """
    try:
        conteggi, sotto_soglia = controlla_e_notifica()
        return jsonify({
            'sotto_soglia': sotto_soglia,
            'email_inviata': len(sotto_soglia) > 0
        })
    except Exception as e:
        print(f'Errore invio notifica: {e}')
        return jsonify({'error': 'Errore interno del server'}), 500
